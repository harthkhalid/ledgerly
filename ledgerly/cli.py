"""Click CLI entry point for Ledgerly."""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import click
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart

from ledgerly.engine import names as N
from ledgerly.engine.workbook import (
    build_workbook,
    count_charts,
    count_formulas,
    save_workbook,
)


def _parse_month(month: str) -> tuple[datetime, datetime]:
    try:
        year, mon = map(int, month.split("-"))
        since = datetime(year, mon, 1)
        if mon == 12:
            until = datetime(year + 1, 1, 1)
        else:
            until = datetime(year, mon + 1, 1)
        return since, until
    except Exception as exc:
        raise click.BadParameter(f"Invalid month '{month}', expected YYYY-MM") from exc


def _load_source(source: str):
    if source == "fixture":
        from ledgerly.adapters.fixture import FixtureSource

        return FixtureSource()
    if source == "shopify":
        store = os.environ.get("SHOPIFY_STORE")
        token = os.environ.get("SHOPIFY_ACCESS_TOKEN")
        if not store or not token:
            raise click.ClickException(
                "Shopify source requires SHOPIFY_STORE and SHOPIFY_ACCESS_TOKEN "
                "environment variables. Example:\n"
                "  set SHOPIFY_STORE=my-store\n"
                "  set SHOPIFY_ACCESS_TOKEN=shpat_...\n"
                "Or use --source fixture for offline generation."
            )
        from ledgerly.adapters.shopify import ShopifySource

        return ShopifySource(store=store, access_token=token)
    raise click.BadParameter(f"Unknown source '{source}'")


def _gather(source, month: str):
    since, until = _parse_month(month)
    products, variants = source.fetch_products()
    inventory = source.fetch_inventory()
    # Prefer full history when available (fixture); else month window
    if hasattr(source, "fetch_all_orders"):
        all_orders = source.fetch_all_orders()
    else:
        # For Shopify, pull a 6-month lookback for cohorts/velocity
        lookback_start = datetime(since.year, since.month, 1)
        # go back 5 months
        m = since.month - 5
        y = since.year
        while m <= 0:
            m += 12
            y -= 1
        lookback_start = datetime(y, m, 1)
        all_orders = source.fetch_orders(lookback_start, until)

    customers = []
    if hasattr(source, "fetch_customers"):
        customers = source.fetch_customers()
    else:
        # Derive minimal customer stubs from orders for Shopify path
        from ledgerly.models import Customer

        seen = {}
        for o in all_orders:
            if o.customer_id not in seen:
                seen[o.customer_id] = Customer(
                    id=o.customer_id,
                    created_at=o.created_at,
                    email_hash="unknown",
                )
            else:
                if o.created_at < seen[o.customer_id].created_at:
                    seen[o.customer_id] = Customer(
                        id=o.customer_id,
                        created_at=o.created_at,
                        email_hash="unknown",
                    )
        customers = list(seen.values())

    return {
        "orders": all_orders,
        "products": products,
        "variants": variants,
        "inventory": inventory,
        "customers": customers,
        "all_orders": all_orders,
        "report_month": month,
    }


@click.group()
@click.version_option(package_name="ledgerly")
def main() -> None:
    """Ledgerly — Shopify operations workbook engine."""


@main.command("generate")
@click.option("--month", required=True, help="Report month as YYYY-MM")
@click.option(
    "--source",
    type=click.Choice(["fixture", "shopify"]),
    default="fixture",
    show_default=True,
    help="Data source adapter",
)
@click.option(
    "--out",
    "out_path",
    type=click.Path(dir_okay=False, path_type=Path),
    default=None,
    help="Output .xlsx path",
)
@click.option(
    "--no-template",
    is_flag=True,
    default=False,
    help="Build from scratch instead of injecting into brand_report.xlsx",
)
def generate(month: str, source: str, out_path: Path | None, no_template: bool) -> None:
    """Generate a fully editable multi-tab Excel operations report."""
    _parse_month(month)  # validate early
    src = _load_source(source)
    data = _gather(src, month)

    if out_path is None:
        out_path = Path(f"ledgerly_report_{month}.xlsx")

    mode = "from-scratch" if no_template else "template"
    if no_template:
        wb = build_workbook(**data)
    else:
        template = (
            Path(__file__).resolve().parent / "templates" / "brand_report.xlsx"
        )
        if not template.exists():
            click.echo(
                f"Template not found at {template}; falling back to from-scratch.",
                err=True,
            )
            wb = build_workbook(**data)
            mode = "from-scratch (template missing)"
        else:
            from ledgerly.engine.template_mode import inject_into_template

            # Template mode injects month orders into Data; also emit full workbook
            # alongside? Spec: template is default generation path.
            # For a complete ops report the from-scratch workbook is the analysis
            # deliverable; template mode produces the branded Summary+Data pack.
            # When --no-template is absent we inject into template AND the user
            # gets the branded file. For full 6-tab analysis use --no-template.
            #
            # Re-read spec: "Default generation path" is template mode.
            # "From-scratch generation is the fallback when --no-template is passed."
            # So default = template injection output.
            month_orders = [
                o
                for o in data["all_orders"]
                if o.created_at.strftime("%Y-%m") == month
            ]
            wb = inject_into_template(
                orders=month_orders or data["all_orders"],
                products=data["products"],
                variants=data["variants"],
                report_month=month,
                template_path=template,
            )

    save_workbook(wb, out_path)

    tabs = [s for s in wb.sheetnames]
    formulas = count_formulas(wb)
    named = len(list(wb.defined_names.keys()))
    charts = count_charts(wb)

    click.echo("Ledgerly report generated successfully.")
    click.echo(f"  Mode:          {mode}")
    click.echo(f"  Source:        {source}")
    click.echo(f"  Month:         {month}")
    click.echo(f"  Tabs:          {', '.join(tabs)}")
    click.echo(f"  Formulas:      {formulas}")
    click.echo(f"  Named ranges:  {named}")
    click.echo(f"  Charts:        {charts}")
    click.echo(f"  Output:        {out_path.resolve()}")


@main.command("validate")
@click.argument("report", type=click.Path(exists=True, dir_okay=False, path_type=Path))
def validate(report: Path) -> None:
    """Reopen a workbook and re-run the assertion suite against it."""
    wb = load_workbook(report)
    errors: list[str] = []

    # Named ranges (full workbook) or template subset
    present = set(wb.defined_names.keys())
    control_names = set(N.ALL_CONTROL_NAMES)
    if control_names <= present:
        for name in N.ALL_CONTROL_NAMES:
            dn = wb.defined_names[name]
            if "Controls" not in (dn.attr_text or ""):
                errors.append(f"Named range {name} does not resolve to Controls")
    else:
        missing = control_names - present
        errors.append(f"Missing control named ranges: {', '.join(sorted(missing))}")

    # If full analysis workbook, check formulas / charts / CF
    if "SKU Margin" in wb.sheetnames:
        g3 = wb["SKU Margin"]["G3"].value
        if not (isinstance(g3, str) and g3.startswith("=") and "ShippingCostPerOrder" in g3):
            errors.append("SKU Margin G3 must be a formula referencing ShippingCostPerOrder")
        if "RawData" in wb.sheetnames:
            k3 = wb["RawData"]["K3"].value
            if not (isinstance(k3, str) and k3.startswith("=")):
                errors.append("RawData K3 (net_revenue) must be a formula")
        if "Dashboard" in wb.sheetnames:
            dash = wb["Dashboard"]
            for row in range(4, 8):
                val = dash.cell(row=row, column=2).value
                if not (isinstance(val, str) and val.startswith("=")):
                    errors.append(f"Dashboard B{row} must be a formula")
            charts = dash._charts
            lines = sum(1 for c in charts if isinstance(c, LineChart))
            bars = sum(1 for c in charts if isinstance(c, BarChart))
            pies = sum(1 for c in charts if isinstance(c, PieChart))
            if lines != 1 or bars != 1 or pies != 1:
                errors.append(
                    f"Dashboard charts expected 1/1/1 Line/Bar/Pie, got {lines}/{bars}/{pies}"
                )

    if errors:
        click.echo("VALIDATION FAILED", err=True)
        for e in errors:
            click.echo(f"  - {e}", err=True)
        sys.exit(1)

    click.echo(f"VALIDATION PASSED: {report}")
    click.echo(f"  Sheets: {', '.join(wb.sheetnames)}")
    click.echo(f"  Named ranges: {len(present)}")
    click.echo(f"  Formulas: {count_formulas(wb)}")


if __name__ == "__main__":
    main()
