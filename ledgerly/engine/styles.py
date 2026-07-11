"""Central style registry — fonts, fills, borders, number formats."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Fonts
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1A1A2E")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=11, bold=False, color="333333")
INPUT_FONT = Font(name="Calibri", size=11, bold=True, color="1A1A2E")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="1A1A2E")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True, color="1A1A2E")

# Fills
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF9E3", end_color="FFF9E3", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F5F5F8", end_color="F5F5F8", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")
KPI_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

# Fonts for CF
RED_FONT = Font(name="Calibri", size=11, color="B71C1C", bold=True)
AMBER_FONT = Font(name="Calibri", size=11, color="E65100", bold=True)
GREEN_FONT = Font(name="Calibri", size=11, color="1B5E20", bold=True)

# Borders
THIN_SIDE = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
DOUBLE_TOP = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=Side(style="double", color="1A1A2E"),
    bottom=THIN_SIDE,
)

# Number formats
CURRENCY_FMT = '$#,##0.00'
PERCENT_FMT = "0.0%"
INT_FMT = "#,##0"
DATE_FMT = "YYYY-MM-DD"
DECIMAL_FMT = "0.00"

# Alignment
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def apply_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    """Style a header row across columns start_col..end_col inclusive."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER


def apply_title(ws, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.font = TITLE_FONT


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
