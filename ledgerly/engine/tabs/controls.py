"""Controls tab — editable business assumptions with defined names."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ledgerly.engine import names as N
from ledgerly.engine import styles as S


def build_controls(wb: Workbook, report_month: str) -> Worksheet:
    ws = wb.create_sheet("Controls", 0)
    S.apply_title(ws, "A1", "Ledgerly — Business Assumptions")
    ws.merge_cells("A1:B1")

    ws["A2"] = "Assumption"
    ws["B2"] = "Value"
    S.apply_header_row(ws, 2, 1, 2)

    defaults = dict(N.DEFAULTS)
    defaults[N.REPORT_MONTH] = report_month

    for name in N.ALL_CONTROL_NAMES:
        row = N.CONTROL_ROWS[name]
        label_cell = ws.cell(row=row, column=1, value=N.CONTROL_LABELS[name])
        label_cell.font = S.LABEL_FONT
        label_cell.border = S.THIN_BORDER
        label_cell.alignment = S.LEFT

        value_cell = ws.cell(row=row, column=2, value=defaults[name])
        value_cell.font = S.INPUT_FONT
        value_cell.fill = S.INPUT_FILL
        value_cell.border = S.THIN_BORDER
        value_cell.alignment = S.RIGHT

        if name in (
            N.SHIPPING_COST_PER_ORDER,
            N.PACKAGING_COST_PER_ORDER,
            N.MONTHLY_AD_SPEND,
        ):
            value_cell.number_format = S.CURRENCY_FMT
        elif name in (N.PAYMENT_PROCESSING_PCT, N.MARGIN_FLOOR_PCT):
            value_cell.number_format = S.PERCENT_FMT
        elif name in (N.LEAD_TIME_DAYS, N.SAFETY_STOCK_DAYS):
            value_cell.number_format = S.INT_FMT

    N.register_control_names(wb)

    ws.freeze_panes = "A3"
    S.set_col_widths(ws, {"A": 32, "B": 18})
    ws.print_area = "A1:B10"
    return ws
