"""Dashboard tab — KPI block + native LineChart, BarChart, PieChart."""

from __future__ import annotations

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ledgerly.engine import names as N
from ledgerly.engine import styles as S


def _month_labels(report_month: str, n: int = 6) -> list[str]:
    year, month = map(int, report_month.split("-"))
    labels = []
    y, m = year, month
    for _ in range(n):
        labels.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(labels))


def build_dashboard(
    wb: Workbook,
    report_month: str,
    raw_last_row: int,
    sku_count: int,
    inv_count: int,
) -> Worksheet:
    ws = wb.create_sheet("Dashboard", 0)  # front
    # Re-order: Dashboard should be first visible after Controls... Spec lists
    # Controls as tab 1. Keep Controls at index 0; move Dashboard after build.
    S.apply_title(ws, "A1", "Ledgerly Operations Dashboard")
    ws.merge_cells("A1:F1")

    sku_first, sku_last = 3, 2 + sku_count
    inv_first, inv_last = 3, 2 + inv_count

    # --- KPI block ---
    ws["A3"] = "KPI Summary"
    ws["A3"].font = S.TOTAL_FONT

    kpis = [
        (4, "Total Revenue", f"='SKU Margin'!D{sku_last + 1}"),
        (5, "Blended Margin %", f"='SKU Margin'!L{sku_last + 1}"),
        (
            6,
            "SKUs Below Margin Floor",
            f'=COUNTIF(\'SKU Margin\'!L{sku_first}:L{sku_last},"<"&{N.MARGIN_FLOOR_PCT})',
        ),
        (
            7,
            "REORDER NOW SKUs",
            f'=COUNTIF(Inventory!G{inv_first}:G{inv_last},"REORDER NOW")',
        ),
    ]
    for row, label, formula in kpis:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = S.KPI_LABEL_FONT
        lc.fill = S.KPI_FILL
        lc.border = S.THIN_BORDER
        vc = ws.cell(row=row, column=2, value=formula)
        vc.font = S.KPI_VALUE_FONT
        vc.fill = S.KPI_FILL
        vc.border = S.THIN_BORDER
        if "Margin %" in label:
            vc.number_format = S.PERCENT_FMT
        elif "Revenue" in label:
            vc.number_format = S.CURRENCY_FMT
        else:
            vc.number_format = S.INT_FMT

    # CF: margin-floor-count red when > 0
    ws.conditional_formatting.add(
        "B6",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=S.RED_FILL,
            font=S.RED_FONT,
        ),
    )

    # --- Hidden helper block for monthly charts (columns AA+) ---
    months = _month_labels(report_month, 6)
    helper_start_row = 3
    ws.cell(row=2, column=27, value="Month")  # AA
    ws.cell(row=2, column=28, value="Net Revenue")  # AB
    ws.cell(row=2, column=29, value="Contribution Profit")  # AC

    raw_date = f"RawData!$B$3:$B${raw_last_row}"
    raw_rev = f"RawData!$K$3:$K${raw_last_row}"

    for i, month in enumerate(months):
        r = helper_start_row + i
        ws.cell(row=r, column=27, value=month)
        # Monthly net revenue via SUMPRODUCT
        ws.cell(
            row=r,
            column=28,
            value=(
                f'=SUMPRODUCT((TEXT({raw_date},"YYYY-MM")=AA{r})*({raw_rev}))'
            ),
        ).number_format = S.CURRENCY_FMT
        # Contribution profit approximation: revenue * blended margin from totals
        # Use SKU Margin total margin % applied to monthly revenue for chart series
        ws.cell(
            row=r,
            column=29,
            value=f"=AB{r}*'SKU Margin'!L{sku_last + 1}",
        ).number_format = S.CURRENCY_FMT

    helper_end = helper_start_row + len(months) - 1

    # Hide helper columns visually by grouping far right; leave values for charts
    ws.column_dimensions["AA"].hidden = True
    ws.column_dimensions["AB"].hidden = True
    ws.column_dimensions["AC"].hidden = True

    # --- LineChart: monthly net revenue vs contribution profit ---
    line = LineChart()
    line.title = "Monthly Revenue vs Contribution Profit"
    line.style = 10
    line.y_axis.title = "Amount ($)"
    line.x_axis.title = "Month"
    line.height = 10
    line.width = 15
    data = Reference(ws, min_col=28, min_row=2, max_col=29, max_row=helper_end)
    cats = Reference(ws, min_col=27, min_row=helper_start_row, max_row=helper_end)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "A10")

    # --- BarChart: top 10 SKUs by contribution $ ---
    # Helper block for top 10: copy formulas referencing SKU Margin sorted isn't
    # possible live without SORT. We build a static top-10 snapshot of SKU labels
    # at generation time is wrong per "formulas" ethos. Instead chart the full
    # SKU Margin contribution column and let Excel show all — spec says top 10.
    # Approach: helper block AE/AF with INDEX/LARGE formulas for top 10.
    ws.cell(row=2, column=31, value="Top SKU")  # AE
    ws.cell(row=2, column=32, value="Contribution $")  # AF
    for i in range(1, 11):
        r = 2 + i
        # Rank-based pull from SKU Margin
        ws.cell(
            row=r,
            column=32,
            value=(
                f"=IFERROR(LARGE('SKU Margin'!$K${sku_first}:$K${sku_last},{i}),0)"
            ),
        ).number_format = S.CURRENCY_FMT
        ws.cell(
            row=r,
            column=31,
            value=(
                f"=IFERROR(INDEX('SKU Margin'!$A${sku_first}:$A${sku_last},"
                f"MATCH(AF{r},'SKU Margin'!$K${sku_first}:$K${sku_last},0)),\"\")"
            ),
        )
    ws.column_dimensions["AE"].hidden = True
    ws.column_dimensions["AF"].hidden = True

    bar = BarChart()
    bar.type = "col"
    bar.title = "Top 10 SKUs by Contribution Margin $"
    bar.style = 10
    bar.y_axis.title = "Contribution ($)"
    bar.x_axis.title = "SKU"
    bar.height = 10
    bar.width = 15
    bar_data = Reference(ws, min_col=32, min_row=2, max_row=12)
    bar_cats = Reference(ws, min_col=31, min_row=3, max_row=12)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    ws.add_chart(bar, "I10")

    # --- PieChart: revenue share by product_type ---
    # Helper: unique product types with SUMIF
    types = ["Eau de Parfum", "Eau de Toilette", "Gift Set", "Home"]
    ws.cell(row=2, column=34, value="Product Type")  # AH
    ws.cell(row=2, column=35, value="Revenue")  # AI
    raw_ptype = f"RawData!$O$3:$O${raw_last_row}"
    for i, ptype in enumerate(types):
        r = 3 + i
        ws.cell(row=r, column=34, value=ptype)
        ws.cell(
            row=r,
            column=35,
            value=f'=SUMIF({raw_ptype},AH{r},{raw_rev})',
        ).number_format = S.CURRENCY_FMT
    ws.column_dimensions["AH"].hidden = True
    ws.column_dimensions["AI"].hidden = True

    pie = PieChart()
    pie.title = "Revenue Share by Product Type"
    pie.style = 10
    pie.height = 10
    pie.width = 12
    pie_data = Reference(ws, min_col=35, min_row=2, max_row=2 + len(types))
    pie_cats = Reference(ws, min_col=34, min_row=3, max_row=2 + len(types))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "A26")

    ws.freeze_panes = "A3"
    S.set_col_widths(ws, {"A": 28, "B": 18, "C": 14, "D": 14, "E": 14, "F": 14})
    ws.print_area = "A1:F40"
    return ws
