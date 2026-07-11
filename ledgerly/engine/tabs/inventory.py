"""Inventory tab — days of cover, reorder dates, status with CF rules."""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ledgerly.engine import names as N
from ledgerly.engine import styles as S
from ledgerly.models import Product, Variant

HEADERS = [
    "SKU",
    "Product",
    "Available",
    "Daily Velocity",
    "Days of Cover",
    "Reorder By",
    "Status",
]


def build_inventory(
    wb: Workbook,
    variants: list[Variant],
    products: list[Product],
    inv_last_row: int,
) -> Worksheet:
    ws = wb.create_sheet("Inventory", 2)
    S.apply_title(ws, "A1", "Inventory & Reorder Planning")
    ws.merge_cells("A1:G1")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = S.HEADER_FONT
        cell.fill = S.HEADER_FILL
        cell.border = S.THIN_BORDER
        cell.alignment = S.CENTER

    product_by_id = {p.id: p for p in products}
    # RawInventory: A=sku, B=title, C=available, D=unit_cost, E=trailing90
    raw_sku = f"RawInventory!$A$3:$A${inv_last_row}"
    raw_avail = f"RawInventory!$C$3:$C${inv_last_row}"
    raw_t90 = f"RawInventory!$E$3:$E${inv_last_row}"

    sorted_variants = sorted(variants, key=lambda v: v.sku)
    first_data = 3
    n = len(sorted_variants)
    last_data = first_data + n - 1 if n else first_data

    for i, variant in enumerate(sorted_variants):
        row = first_data + i
        product = product_by_id.get(variant.product_id)
        title = product.title if product else variant.title

        ws.cell(row=row, column=1, value=variant.sku).border = S.THIN_BORDER
        ws.cell(row=row, column=2, value=title).border = S.THIN_BORDER

        c = ws.cell(
            row=row,
            column=3,
            value=f'=SUMIF({raw_sku},$A{row},{raw_avail})',
        )
        c.number_format = S.INT_FMT
        c.border = S.THIN_BORDER

        # Daily velocity = trailing90 / 90
        c = ws.cell(
            row=row,
            column=4,
            value=f'=SUMIF({raw_sku},$A{row},{raw_t90})/90',
        )
        c.number_format = S.DECIMAL_FMT
        c.border = S.THIN_BORDER

        # Days of cover
        c = ws.cell(
            row=row,
            column=5,
            value=f"=IFERROR(C{row}/D{row},999)",
        )
        c.number_format = S.DECIMAL_FMT
        c.border = S.THIN_BORDER

        # Reorder-by date
        c = ws.cell(
            row=row,
            column=6,
            value=(
                f"=IF(E{row}-{N.LEAD_TIME_DAYS}-{N.SAFETY_STOCK_DAYS}<=0,"
                f"TODAY(),TODAY()+E{row}-{N.LEAD_TIME_DAYS}-{N.SAFETY_STOCK_DAYS})"
            ),
        )
        c.number_format = S.DATE_FMT
        c.border = S.THIN_BORDER

        # Status
        c = ws.cell(
            row=row,
            column=7,
            value=(
                f'=IF(E{row}<{N.LEAD_TIME_DAYS},"REORDER NOW",'
                f'IF(E{row}<{N.LEAD_TIME_DAYS}+{N.SAFETY_STOCK_DAYS},"REORDER SOON","OK"))'
            ),
        )
        c.border = S.THIN_BORDER
        c.alignment = S.CENTER

    if n > 0:
        status_range = f"G{first_data}:G{last_data}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"REORDER NOW"'],
                fill=S.RED_FILL,
                font=S.RED_FONT,
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"REORDER SOON"'],
                fill=S.AMBER_FILL,
                font=S.AMBER_FONT,
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"OK"'],
                fill=S.GREEN_FILL,
                font=S.GREEN_FONT,
            ),
        )

    ws.freeze_panes = "A3"
    S.set_col_widths(
        ws,
        {"A": 14, "B": 20, "C": 12, "D": 14, "E": 14, "F": 14, "G": 14},
    )
    ws.print_area = f"A1:G{last_data}"
    return ws
