"""RawData / RawInventory hidden sheets — flat line table + inventory snapshot.

Trailing-90-day units sold on RawInventory is a static value computed at generation
time. Excel cannot maintain a rolling 90-day window against absolute dates without
volatile TODAY()-based helpers that shift every open; velocity is intentionally a
report-generation snapshot so Days-of-Cover stays stable for the report month.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal

from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ledgerly.engine import styles as S
from ledgerly.models import Customer, InventoryLevel, Order, Product, Variant

# Column layout for tblLines (1-indexed)
COL = {
    "order_id": 1,       # A
    "date": 2,           # B
    "customer_id": 3,    # C
    "first_order_date": 4,  # D — static lookup
    "cohort_month": 5,   # E — formula TEXT
    "sku": 6,            # F
    "product_title": 7,  # G
    "quantity": 8,       # H
    "unit_price": 9,     # I
    "discount": 10,      # J
    "net_revenue": 11,   # K — formula
    "unit_cost": 12,     # L
    "cogs": 13,          # M — formula
    "order_fraction": 14,  # N — static 1/n_lines (for shipping allocation)
    "product_type": 15,  # O
}

HEADERS = [
    "order_id",
    "date",
    "customer_id",
    "first_order_date",
    "cohort_month",
    "sku",
    "product_title",
    "quantity",
    "unit_price",
    "discount",
    "net_revenue",
    "unit_cost",
    "cogs",
    "order_fraction",
    "product_type",
]

INV_HEADERS = ["sku", "product_title", "available_qty", "unit_cost", "trailing90_units", "product_type"]


def _customer_first_order(
    orders: list[Order], customers: list[Customer]
) -> dict[str, datetime]:
    """Earliest order date per customer; fall back to customer.created_at."""
    first: dict[str, datetime] = {c.id: c.created_at for c in customers}
    for o in orders:
        prev = first.get(o.customer_id)
        if prev is None or o.created_at < prev:
            first[o.customer_id] = o.created_at
    return first


def build_raw_data(
    wb: Workbook,
    orders: list[Order],
    products: list[Product],
    variants: list[Variant],
    customers: list[Customer],
    inventory: list[InventoryLevel],
    report_end: datetime,
) -> tuple[Worksheet, Worksheet, int]:
    """Build RawData + RawInventory. Returns (raw_ws, inv_ws, line_count)."""
    product_by_id = {p.id: p for p in products}
    variant_by_id = {v.id: v for v in variants}
    variant_by_sku = {v.sku: v for v in variants}
    inv_by_item = {i.inventory_item_id: i for i in inventory}
    first_order = _customer_first_order(orders, customers)

    ws = wb.create_sheet("RawData")
    S.apply_title(ws, "A1", "Raw Order Lines")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = S.HEADER_FONT
        cell.fill = S.HEADER_FILL
        cell.border = S.THIN_BORDER

    # Sort orders for stable output
    sorted_orders = sorted(orders, key=lambda o: (o.created_at, o.id))
    row = 3
    for order in sorted_orders:
        n_lines = max(len(order.line_items), 1)
        frac = 1.0 / n_lines
        fod = first_order.get(order.customer_id, order.created_at)
        for li in order.line_items:
            variant = variant_by_id.get(li.variant_id) or variant_by_sku.get(li.sku)
            product = product_by_id.get(variant.product_id) if variant else None
            unit_cost = variant.cost if variant else Decimal("0")
            title = product.title if product else (variant.title if variant else li.sku)
            ptype = product.product_type if product else ""

            ws.cell(row=row, column=COL["order_id"], value=order.id)
            date_cell = ws.cell(row=row, column=COL["date"], value=order.created_at)
            date_cell.number_format = S.DATE_FMT
            ws.cell(row=row, column=COL["customer_id"], value=order.customer_id)
            fod_cell = ws.cell(row=row, column=COL["first_order_date"], value=fod)
            fod_cell.number_format = S.DATE_FMT
            # cohort_month formula from first-order lookup column
            ws.cell(
                row=row,
                column=COL["cohort_month"],
                value=f'=TEXT(D{row},"YYYY-MM")',
            )
            ws.cell(row=row, column=COL["sku"], value=li.sku)
            ws.cell(row=row, column=COL["product_title"], value=title)
            ws.cell(row=row, column=COL["quantity"], value=li.quantity)
            price_cell = ws.cell(
                row=row, column=COL["unit_price"], value=float(li.unit_price)
            )
            price_cell.number_format = S.CURRENCY_FMT
            disc_cell = ws.cell(
                row=row, column=COL["discount"], value=float(li.discount_allocated)
            )
            disc_cell.number_format = S.CURRENCY_FMT
            # net_revenue = qty * price - discount
            net_cell = ws.cell(
                row=row,
                column=COL["net_revenue"],
                value=f"=H{row}*I{row}-J{row}",
            )
            net_cell.number_format = S.CURRENCY_FMT
            cost_cell = ws.cell(
                row=row, column=COL["unit_cost"], value=float(unit_cost)
            )
            cost_cell.number_format = S.CURRENCY_FMT
            cogs_cell = ws.cell(
                row=row, column=COL["cogs"], value=f"=H{row}*L{row}"
            )
            cogs_cell.number_format = S.CURRENCY_FMT
            ws.cell(row=row, column=COL["order_fraction"], value=frac)
            ws.cell(row=row, column=COL["product_type"], value=ptype)
            row += 1

    line_count = row - 3
    last_data_row = max(row - 1, 3)

    if line_count > 0:
        table = Table(
            displayName="tblLines",
            ref=f"A2:O{last_data_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.freeze_panes = "A3"
    ws.sheet_state = "hidden"
    S.set_col_widths(
        ws,
        {
            "A": 14, "B": 12, "C": 14, "D": 14, "E": 12,
            "F": 14, "G": 22, "H": 10, "I": 12, "J": 10,
            "K": 12, "L": 10, "M": 10, "N": 12, "O": 14,
        },
    )
    ws.print_area = f"A1:O{last_data_row}"

    # --- RawInventory ---
    # Trailing 90-day units: static snapshot (see module docstring).
    window_start = report_end - timedelta(days=90)
    units_90: dict[str, int] = defaultdict(int)
    for order in orders:
        if window_start <= order.created_at < report_end:
            for li in order.line_items:
                units_90[li.sku] += li.quantity

    inv_ws = wb.create_sheet("RawInventory")
    S.apply_title(inv_ws, "A1", "Raw Inventory Snapshot")
    for col, header in enumerate(INV_HEADERS, start=1):
        cell = inv_ws.cell(row=2, column=col, value=header)
        cell.font = S.HEADER_FONT
        cell.fill = S.HEADER_FILL
        cell.border = S.THIN_BORDER

    inv_row = 3
    for variant in sorted(variants, key=lambda v: v.sku):
        product = product_by_id.get(variant.product_id)
        inv = inv_by_item.get(variant.inventory_item_id)
        avail = inv.available if inv else 0
        inv_ws.cell(row=inv_row, column=1, value=variant.sku)
        inv_ws.cell(
            row=inv_row, column=2, value=product.title if product else variant.title
        )
        inv_ws.cell(row=inv_row, column=3, value=avail)
        cost_c = inv_ws.cell(row=inv_row, column=4, value=float(variant.cost))
        cost_c.number_format = S.CURRENCY_FMT
        inv_ws.cell(row=inv_row, column=5, value=units_90.get(variant.sku, 0))
        inv_ws.cell(
            row=inv_row,
            column=6,
            value=product.product_type if product else "",
        )
        inv_row += 1

    inv_last = max(inv_row - 1, 3)
    inv_ws.freeze_panes = "A3"
    inv_ws.sheet_state = "hidden"
    S.set_col_widths(
        inv_ws, {"A": 14, "B": 22, "C": 14, "D": 12, "E": 16, "F": 14}
    )
    inv_ws.print_area = f"A1:F{inv_last}"

    return ws, inv_ws, line_count
