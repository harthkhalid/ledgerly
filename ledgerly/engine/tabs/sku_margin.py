"""SKU Margin tab — contribution margin with live SUMIFS and named-range allocations."""

from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ledgerly.engine import names as N
from ledgerly.engine import styles as S
from ledgerly.models import Product, Variant

# Column letters on SKU Margin (data starts row 3)
# A SKU | B Product | C Units | D Gross Rev | E COGS | F Order Share
# G Alloc Shipping | H Alloc Packaging | I Processing | J Ad Spend
# K Contrib $ | L Contrib %


HEADERS = [
    "SKU",
    "Product",
    "Units Sold",
    "Gross Revenue",
    "COGS",
    "Order Share",
    "Allocated Shipping",
    "Allocated Packaging",
    "Processing Fees",
    "Ad Spend Alloc",
    "Contribution $",
    "Contribution %",
]


def build_sku_margin(
    wb: Workbook,
    variants: list[Variant],
    products: list[Product],
    raw_last_row: int,
) -> Worksheet:
    ws = wb.create_sheet("SKU Margin", 1)
    S.apply_title(ws, "A1", "SKU Contribution Margin")
    ws.merge_cells("A1:J1")
    ws["K1"] = "Margin Floor"
    ws["K1"].font = S.KPI_LABEL_FONT
    ws["L1"] = f"={N.MARGIN_FLOOR_PCT}"
    ws["L1"].number_format = S.PERCENT_FMT
    ws["L1"].font = S.INPUT_FONT

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = S.HEADER_FONT
        cell.fill = S.HEADER_FILL
        cell.border = S.THIN_BORDER
        cell.alignment = S.CENTER

    product_by_id = {p.id: p for p in products}
    # RawData columns: F=sku, H=qty, K=net_revenue, M=cogs, N=order_fraction
    raw_sku = f"RawData!$F$3:$F${raw_last_row}"
    raw_qty = f"RawData!$H$3:$H${raw_last_row}"
    raw_rev = f"RawData!$K$3:$K${raw_last_row}"
    raw_cogs = f"RawData!$M$3:$M${raw_last_row}"
    raw_frac = f"RawData!$N$3:$N${raw_last_row}"

    sorted_variants = sorted(variants, key=lambda v: v.sku)
    n = len(sorted_variants)
    first_data = 3
    last_data = first_data + n - 1 if n else first_data

    for i, variant in enumerate(sorted_variants):
        row = first_data + i
        product = product_by_id.get(variant.product_id)
        title = product.title if product else variant.title

        ws.cell(row=row, column=1, value=variant.sku).border = S.THIN_BORDER
        ws.cell(row=row, column=2, value=title).border = S.THIN_BORDER

        # Units sold
        c = ws.cell(
            row=row,
            column=3,
            value=f'=SUMIFS({raw_qty},{raw_sku},$A{row})',
        )
        c.number_format = S.INT_FMT
        c.border = S.THIN_BORDER

        # Gross revenue
        c = ws.cell(
            row=row,
            column=4,
            value=f'=SUMIFS({raw_rev},{raw_sku},$A{row})',
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # COGS
        c = ws.cell(
            row=row,
            column=5,
            value=f'=SUMIFS({raw_cogs},{raw_sku},$A{row})',
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Order share (fractional order count)
        c = ws.cell(
            row=row,
            column=6,
            value=f'=SUMIFS({raw_frac},{raw_sku},$A{row})',
        )
        c.number_format = S.DECIMAL_FMT
        c.border = S.THIN_BORDER

        # Allocated shipping
        c = ws.cell(
            row=row,
            column=7,
            value=f"=F{row}*{N.SHIPPING_COST_PER_ORDER}",
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Allocated packaging
        c = ws.cell(
            row=row,
            column=8,
            value=f"=F{row}*{N.PACKAGING_COST_PER_ORDER}",
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Processing fees
        c = ws.cell(
            row=row,
            column=9,
            value=f"=D{row}*{N.PAYMENT_PROCESSING_PCT}",
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Ad spend allocation (revenue-weighted)
        c = ws.cell(
            row=row,
            column=10,
            value=f"=IF(SUM($D${first_data}:$D${last_data})=0,0,{N.MONTHLY_AD_SPEND}*(D{row}/SUM($D${first_data}:$D${last_data})))",
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Contribution $
        c = ws.cell(
            row=row,
            column=11,
            value=f"=D{row}-E{row}-G{row}-H{row}-I{row}-J{row}",
        )
        c.number_format = S.CURRENCY_FMT
        c.border = S.THIN_BORDER

        # Contribution %
        c = ws.cell(
            row=row,
            column=12,
            value=f"=IF(D{row}=0,0,K{row}/D{row})",
        )
        c.number_format = S.PERCENT_FMT
        c.border = S.THIN_BORDER

    # Totals row
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = S.TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = S.TOTAL_FILL
    ws.cell(row=total_row, column=1).border = S.DOUBLE_TOP
    ws.cell(row=total_row, column=2).border = S.DOUBLE_TOP
    ws.cell(row=total_row, column=2).fill = S.TOTAL_FILL

    for col, fmt in [
        (3, S.INT_FMT),
        (4, S.CURRENCY_FMT),
        (5, S.CURRENCY_FMT),
        (6, S.DECIMAL_FMT),
        (7, S.CURRENCY_FMT),
        (8, S.CURRENCY_FMT),
        (9, S.CURRENCY_FMT),
        (10, S.CURRENCY_FMT),
        (11, S.CURRENCY_FMT),
    ]:
        letter = chr(64 + col)
        c = ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({letter}{first_data}:{letter}{last_data})",
        )
        c.number_format = fmt
        c.font = S.TOTAL_FONT
        c.fill = S.TOTAL_FILL
        c.border = S.DOUBLE_TOP

    # Blended margin %
    c = ws.cell(
        row=total_row,
        column=12,
        value=f"=IF(D{total_row}=0,0,K{total_row}/D{total_row})",
    )
    c.number_format = S.PERCENT_FMT
    c.font = S.TOTAL_FONT
    c.fill = S.TOTAL_FILL
    c.border = S.DOUBLE_TOP

    # Conditional formatting: entire data range, flag when margin % < MarginFloorPct
    if n > 0:
        # Formula relative to top-left of range (A3); L3 is the margin % of first row
        formula = f"$L3<{N.MARGIN_FLOOR_PCT}"
        ws.conditional_formatting.add(
            f"A{first_data}:L{last_data}",
            FormulaRule(
                formula=[formula],
                fill=S.RED_FILL,
                font=S.RED_FONT,
            ),
        )

    ws.freeze_panes = "A3"
    S.set_col_widths(
        ws,
        {
            "A": 14, "B": 20, "C": 12, "D": 14, "E": 12, "F": 12,
            "G": 16, "H": 16, "I": 14, "J": 14, "K": 14, "L": 14,
        },
    )
    ws.print_area = f"A1:L{total_row}"
    return ws
