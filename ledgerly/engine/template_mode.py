"""Template-mode injection — flagship path that preserves every template style.

IMPORTANT: Always load with openpyxl.load_workbook(path) using the defaults
(read_only=False, data_only=False). data_only=True would replace every formula
with its last-cached value (or None), permanently destroying the live formula
graph that makes the workbook useful. Never use data_only for injection.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from ledgerly.engine import names as N
from ledgerly.models import Order, Product, Variant

# Names the template must provide for injection to proceed
REQUIRED_TEMPLATE_NAMES = (
    *N.ALL_CONTROL_NAMES,
    "LogoPlaceholder",
    "ReportTitleCell",
    "DataTableAnchor",
)

DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parent.parent / "templates" / "brand_report.xlsx"
)


class TemplateInjectionError(ValueError):
    """Raised when the template is missing required named ranges."""


def _resolve_name(wb: Workbook, name: str) -> tuple[str, str]:
    """Return (sheet_title, cell_coord) for a workbook-level defined name."""
    dn = wb.defined_names[name]
    # attr_text like "Controls!$B$3" or "'Sheet Name'!$A$1"
    destinations = list(dn.destinations)
    if not destinations:
        # Fallback parse attr_text
        attr = dn.attr_text
        sheet, _, coord = attr.replace("'", "").partition("!")
        return sheet, coord.replace("$", "")
    sheet, coord = destinations[0]
    return sheet, coord.replace("$", "")


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    """Copy font/fill/border/number_format/alignment from source_row to target_row."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)


def _extend_table(ws, table_name: str, new_ref: str) -> None:
    if table_name not in ws.tables:
        # openpyxl stores tables in ws.tables dict by name
        # Also check values
        found = None
        for t in ws.tables.values():
            if t.displayName == table_name or t.name == table_name:
                found = t
                break
        if found is None:
            raise TemplateInjectionError(
                f"Template Data sheet is missing Excel Table '{table_name}'"
            )
        found.ref = new_ref
        return
    ws.tables[table_name].ref = new_ref


def inject_into_template(
    *,
    orders: list[Order],
    products: list[Product],
    variants: list[Variant],
    report_month: str,
    template_path: Path | str | None = None,
    control_overrides: dict[str, Any] | None = None,
) -> Workbook:
    """Load the brand template and inject data into designated regions only.

    Does NOT touch any cell, style, formula, column width, or sheet outside the
    designated named ranges and the Data table body.
    """
    path = Path(template_path) if template_path else DEFAULT_TEMPLATE
    if not path.exists():
        raise FileNotFoundError(
            f"Template not found at {path}. Run scripts/make_template.py first."
        )

    # data_only=False is critical — see module docstring.
    wb = load_workbook(path)  # NOT read_only, NOT data_only

    missing = [n for n in REQUIRED_TEMPLATE_NAMES if n not in wb.defined_names]
    if missing:
        raise TemplateInjectionError(
            f"Template is missing required named ranges: {', '.join(sorted(missing))}"
        )

    # --- Inject control values into named-range cells ---
    values = dict(N.DEFAULTS)
    values[N.REPORT_MONTH] = report_month
    if control_overrides:
        values.update(control_overrides)

    for name in N.ALL_CONTROL_NAMES:
        sheet_title, coord = _resolve_name(wb, name)
        wb[sheet_title][coord] = values[name]

    # --- Inject report title (designated named cell only) ---
    title_sheet, title_coord = _resolve_name(wb, "ReportTitleCell")
    wb[title_sheet][title_coord] = f"Ledgerly Brand Report — {report_month}"

    # Logo placeholder intentionally left untouched (style preservation demo)

    # --- Inject rows into Data table ---
    product_by_id = {p.id: p for p in products}
    variant_by_id = {v.id: v for v in variants}
    variant_by_sku = {v.sku: v for v in variants}

    data_ws = wb["Data"]
    # Find style source row (last seed row before overwrite) — row 4 in template
    style_source_row = 4
    max_col = 6

    # Clear existing data rows (keep header row 2); start writing at row 3
    # Remove old body but keep at least the style template row's styles via copy
    rows_out: list[tuple] = []
    for order in sorted(orders, key=lambda o: (o.created_at, o.id)):
        for li in order.line_items:
            variant = variant_by_id.get(li.variant_id) or variant_by_sku.get(li.sku)
            product = product_by_id.get(variant.product_id) if variant else None
            title = product.title if product else (variant.title if variant else li.sku)
            net = float(li.unit_price) * li.quantity - float(li.discount_allocated)
            cost = float(variant.cost) if variant else 0.0
            rows_out.append(
                (
                    li.sku,
                    title,
                    li.quantity,
                    float(li.unit_price),
                    net,
                    cost,
                )
            )

    # Write rows, copying style from the last template body row
    start_row = 3
    for i, row_vals in enumerate(rows_out):
        r = start_row + i
        if r > style_source_row:
            _copy_row_style(data_ws, style_source_row, r, max_col)
        for c, val in enumerate(row_vals, start=1):
            data_ws.cell(row=r, column=c, value=val)

    # Clear any leftover seed cells beyond injected rows
    old_last = 4  # template had rows 3-4
    new_last = start_row + len(rows_out) - 1 if rows_out else start_row
    if new_last < old_last:
        for r in range(new_last + 1, old_last + 1):
            for c in range(1, max_col + 1):
                data_ws.cell(row=r, column=c).value = None

    # Extend / shrink the Excel Table range
    end_row = max(new_last, start_row)
    new_ref = f"A2:F{end_row}"
    _extend_table(data_ws, "tblTemplateData", new_ref)

    return wb


def validate_template_names(path: Path | str) -> list[str]:
    """Return list of missing required names (empty if complete)."""
    wb = load_workbook(path)
    return [n for n in REQUIRED_TEMPLATE_NAMES if n not in wb.defined_names]
