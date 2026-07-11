"""Workbook orchestrator — builds a full from-scratch Ledgerly report.

The engine has ZERO knowledge of Shopify. It consumes pydantic models only.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from ledgerly.engine.tabs.cohort_ltv import build_cohort_ltv
from ledgerly.engine.tabs.controls import build_controls
from ledgerly.engine.tabs.dashboard import build_dashboard
from ledgerly.engine.tabs.inventory import build_inventory
from ledgerly.engine.tabs.raw_data import build_raw_data
from ledgerly.engine.tabs.sku_margin import build_sku_margin
from ledgerly.models import Customer, InventoryLevel, Order, Product, Variant


def build_workbook(
    *,
    orders: list[Order],
    products: list[Product],
    variants: list[Variant],
    inventory: list[InventoryLevel],
    customers: list[Customer],
    report_month: str,
    all_orders: list[Order] | None = None,
) -> Workbook:
    """Build a complete multi-tab workbook with live formulas.

    ``all_orders`` may span a wider window than the report month (needed for
    cohorts and trailing-90 velocity). Defaults to ``orders``.
    """
    year, month = map(int, report_month.split("-"))
    if month == 12:
        report_end = datetime(year + 1, 1, 1)
    else:
        report_end = datetime(year, month + 1, 1)

    source_orders = all_orders if all_orders is not None else orders

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_controls(wb, report_month)
    _, _, line_count = build_raw_data(
        wb,
        orders=source_orders,
        products=products,
        variants=variants,
        customers=customers,
        inventory=inventory,
        report_end=report_end,
    )
    raw_last_row = 2 + max(line_count, 1)
    inv_last_row = 2 + max(len(variants), 1)

    build_sku_margin(wb, variants, products, raw_last_row)
    build_inventory(wb, variants, products, inv_last_row)
    build_cohort_ltv(wb, report_month, raw_last_row)
    build_dashboard(
        wb,
        report_month=report_month,
        raw_last_row=raw_last_row,
        sku_count=len(variants),
        inv_count=len(variants),
    )

    # Sheet order: Controls, Dashboard, SKU Margin, Inventory, Cohort LTV, (hidden raw)
    desired = [
        "Controls",
        "Dashboard",
        "SKU Margin",
        "Inventory",
        "Cohort LTV",
        "RawData",
        "RawInventory",
    ]
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return wb


def save_workbook(wb: Workbook, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def count_formulas(wb: Workbook) -> int:
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n += 1
    return n


def count_charts(wb: Workbook) -> int:
    return sum(len(ws._charts) for ws in wb.worksheets)
