"""Build templates/brand_report.xlsx — reproducible pre-styled injection target.

Run: python scripts/make_template.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from ledgerly.engine import names as N  # noqa: E402

OUT = ROOT / "ledgerly" / "templates" / "brand_report.xlsx"

# Brand styling unique to the template (must survive injection)
BRAND_TITLE_FONT = Font(name="Georgia", size=20, bold=True, color="4A1942")
BRAND_HEADER_FILL = PatternFill(start_color="4A1942", end_color="4A1942", fill_type="solid")
BRAND_HEADER_FONT = Font(name="Georgia", size=11, bold=True, color="FFFFFF")
BRAND_ACCENT_FILL = PatternFill(start_color="F3E5F0", end_color="F3E5F0", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF9E3", end_color="FFF9E3", fill_type="solid")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Extra named ranges for injection regions
INJECT_NAMES = {
    **N.CONTROL_CELLS,
    "LogoPlaceholder": "Summary!$A$2",
    "ReportTitleCell": "Summary!$A$1",
    "DataTableAnchor": "Data!$A$3",
}


def build_template() -> Workbook:
    wb = Workbook()

    # --- Summary sheet (pre-styled, formulas must be preserved) ---
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Ledgerly Brand Report"
    summary["A1"].font = BRAND_TITLE_FONT
    summary.merge_cells("A1:D1")

    summary["A2"] = "[LOGO]"
    summary["A2"].font = Font(name="Georgia", size=10, italic=True, color="9E9E9E")
    summary["A2"].fill = BRAND_ACCENT_FILL

    summary["A4"] = "Executive Summary"
    summary["A4"].font = Font(name="Georgia", size=14, bold=True, color="4A1942")

    summary["A6"] = "Total Line Revenue"
    summary["B6"] = '=SUMIF(Data!A:A,"<>",Data!E:E)'  # sum net_revenue col E where sku present
    summary["B6"].number_format = '$#,##0.00'
    summary["B6"].font = Font(name="Georgia", size=12, bold=True)

    summary["A7"] = "Total Units"
    summary["B7"] = '=SUMIF(Data!A:A,"<>",Data!C:C)'
    summary["B7"].number_format = "#,##0"
    summary["B7"].font = Font(name="Georgia", size=12, bold=True)

    summary["A8"] = "Line Count"
    summary["B8"] = '=COUNTA(Data!A:A)-1'
    summary["B8"].number_format = "#,##0"

    summary["A10"] = "Assumptions mirror"
    summary["A10"].font = Font(name="Georgia", size=11, bold=True, color="4A1942")
    summary["A11"] = "Shipping / order"
    summary["B11"] = f"={N.SHIPPING_COST_PER_ORDER}"
    summary["B11"].number_format = '$#,##0.00'
    summary["A12"] = "Margin floor"
    summary["B12"] = f"={N.MARGIN_FLOOR_PCT}"
    summary["B12"].number_format = "0.0%"

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    summary.freeze_panes = "A5"
    summary.print_area = "A1:D14"

    # --- Controls sheet ---
    controls = wb.create_sheet("Controls")
    controls["A1"] = "Controls"
    controls["A1"].font = BRAND_TITLE_FONT
    controls["A2"] = "Assumption"
    controls["B2"] = "Value"
    for col in (1, 2):
        controls.cell(row=2, column=col).font = BRAND_HEADER_FONT
        controls.cell(row=2, column=col).fill = BRAND_HEADER_FILL
        controls.cell(row=2, column=col).border = THIN

    defaults = dict(N.DEFAULTS)
    for name in N.ALL_CONTROL_NAMES:
        row = N.CONTROL_ROWS[name]
        controls.cell(row=row, column=1, value=N.CONTROL_LABELS[name]).border = THIN
        cell = controls.cell(row=row, column=2, value=defaults[name])
        cell.fill = INPUT_FILL
        cell.border = THIN
        if name in (N.PAYMENT_PROCESSING_PCT, N.MARGIN_FLOOR_PCT):
            cell.number_format = "0.0%"
        elif name in (N.SHIPPING_COST_PER_ORDER, N.PACKAGING_COST_PER_ORDER, N.MONTHLY_AD_SPEND):
            cell.number_format = '$#,##0.00'

    controls.column_dimensions["A"].width = 32
    controls.column_dimensions["B"].width = 16
    controls.freeze_panes = "A3"

    # --- Data sheet with seed table (2 placeholder rows for style copying) ---
    data = wb.create_sheet("Data")
    data["A1"] = "Injected Line Data"
    data["A1"].font = BRAND_TITLE_FONT
    headers = ["sku", "product_title", "quantity", "unit_price", "net_revenue", "unit_cost"]
    for col, h in enumerate(headers, start=1):
        cell = data.cell(row=2, column=col, value=h)
        cell.font = BRAND_HEADER_FONT
        cell.fill = BRAND_HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center")

    # Two seed rows — styles to copy when appending
    seed = [
        ("SEED-001", "Placeholder A", 0, 0.0, 0.0, 0.0),
        ("SEED-002", "Placeholder B", 0, 0.0, 0.0, 0.0),
    ]
    for i, row_vals in enumerate(seed):
        r = 3 + i
        for c, val in enumerate(row_vals, start=1):
            cell = data.cell(row=r, column=c, value=val)
            cell.border = THIN
            cell.font = Font(name="Georgia", size=10, color="4A1942")
            if c >= 4:
                cell.number_format = '$#,##0.00'
            if i == 1:
                cell.fill = BRAND_ACCENT_FILL

    table = Table(displayName="tblTemplateData", ref="A2:F4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
    )
    data.add_table(table)
    for letter, w in zip("ABCDEF", [14, 22, 10, 12, 12, 12]):
        data.column_dimensions[letter].width = w
    data.freeze_panes = "A3"

    # Register all named ranges
    for name, attr in INJECT_NAMES.items():
        wb.defined_names.add(DefinedName(name=name, attr_text=attr))

    return wb


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_template()
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
