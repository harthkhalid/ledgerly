"""Template mode — style preservation and missing-name error path."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

from ledgerly.adapters.fixture import FixtureSource
from ledgerly.engine.template_mode import (
    REQUIRED_TEMPLATE_NAMES,
    TemplateInjectionError,
    inject_into_template,
)
from scripts.make_template import build_template


def _snapshot_cell(cell):
    """Capture style + value identity for preservation asserts."""
    return {
        "value": cell.value,
        "font_name": cell.font.name,
        "font_size": cell.font.size,
        "font_color": cell.font.color.rgb if cell.font.color and cell.font.color.rgb else None,
        "font_bold": cell.font.bold,
        "fill_fg": cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.patternType else None,
        "fill_pattern": cell.fill.patternType,
        "number_format": cell.number_format,
    }


def test_template_injection_preserves_styles_and_summary_formulas(tmp_path):
    # Build template to a temp path
    tmpl_path = tmp_path / "brand_report.xlsx"
    wb_tmpl = build_template()
    wb_tmpl.save(tmpl_path)

    # Snapshot styled cells + Summary formulas BEFORE injection
    before = load_workbook(tmpl_path)
    snap_a2 = _snapshot_cell(before["Summary"]["A2"])  # logo placeholder
    snap_a4 = _snapshot_cell(before["Summary"]["A4"])
    summary_formulas = {
        "B6": before["Summary"]["B6"].value,
        "B7": before["Summary"]["B7"].value,
        "B8": before["Summary"]["B8"].value,
        "B11": before["Summary"]["B11"].value,
        "B12": before["Summary"]["B12"].value,
    }
    # Control default before
    shipping_before = before["Controls"]["B3"].value
    before.close()

    src = FixtureSource()
    products, variants = src.fetch_products()
    orders = src.fetch_orders(datetime(2026, 6, 1), datetime(2026, 7, 1))

    wb = inject_into_template(
        orders=orders[:50],  # keep test fast
        products=products,
        variants=variants,
        report_month="2026-06",
        template_path=tmpl_path,
        control_overrides={"ShippingCostPerOrder": 99.99},
    )
    out = tmp_path / "injected.xlsx"
    wb.save(out)

    after = load_workbook(out)
    # Styles on untouched cells must be byte-identical (field-level)
    assert _snapshot_cell(after["Summary"]["A2"]) == snap_a2
    assert _snapshot_cell(after["Summary"]["A4"]) == snap_a4

    # Pre-existing Summary formulas untouched
    for coord, formula in summary_formulas.items():
        assert after["Summary"][coord].value == formula

    # Injected named-range cells changed
    assert after["Controls"]["B3"].value == 99.99
    assert after["Controls"]["B3"].value != shipping_before

    # Title cell injected
    assert "2026-06" in str(after["Summary"]["A1"].value)

    # Data table extended beyond seed rows
    assert after["Data"]["A3"].value != "SEED-001"
    table = after["Data"].tables["tblTemplateData"]
    # ref should extend past row 4
    assert "F" in table.ref
    end_row = int(table.ref.split(":")[1].replace("F", ""))
    assert end_row > 4


def test_missing_named_range_raises_clear_error(tmp_path):
    tmpl_path = tmp_path / "broken.xlsx"
    wb = build_template()
    # Remove one required name
    del wb.defined_names["MarginFloorPct"]
    wb.save(tmpl_path)

    src = FixtureSource()
    products, variants = src.fetch_products()
    orders = src.fetch_all_orders()[:5]

    with pytest.raises(TemplateInjectionError) as exc_info:
        inject_into_template(
            orders=orders,
            products=products,
            variants=variants,
            report_month="2026-06",
            template_path=tmpl_path,
        )
    msg = str(exc_info.value)
    assert "MarginFloorPct" in msg
    assert "missing required named ranges" in msg.lower()


def test_require_names_lists_every_missing(tmp_path):
    tmpl_path = tmp_path / "broken2.xlsx"
    wb = build_template()
    del wb.defined_names["LeadTimeDays"]
    del wb.defined_names["LogoPlaceholder"]
    wb.save(tmpl_path)

    src = FixtureSource()
    products, variants = src.fetch_products()
    with pytest.raises(TemplateInjectionError) as exc_info:
        inject_into_template(
            orders=[],
            products=products,
            variants=variants,
            report_month="2026-06",
            template_path=tmpl_path,
        )
    msg = str(exc_info.value)
    assert "LeadTimeDays" in msg
    assert "LogoPlaceholder" in msg
